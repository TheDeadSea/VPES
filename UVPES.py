# © [2024] National Electric Vehicle Centre, Land Transport Authority.
# All rights reserved. This code, in full or in part, is the property of the Land Transport Authority.
# No part of this code may be disclosed, reproduced, or distributed without prior written permission.

# Last updated: 21/06/2024
# Ver. 1.3

# Used Vehicle Used Price Extraction System (UVPES)

import os
import time
import random
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import re

# Generate random user agents
def generate_random_user_agent():
    browsers = ['Mozilla/5.0', 'AppleWebKit/537.36', 'Chrome/90.0.4430.93', 'Safari/537.36', 'Firefox/89.0']
    platforms = ['Windows NT 10.0; Win64; x64', 'Macintosh; Intel Mac OS X 10_15_7', 'X11; Ubuntu; Linux x86_64']
    return f"{random.choice(browsers)} ({random.choice(platforms)}) {random.choice(browsers)}"

# Setup
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument(f"user-agent={generate_random_user_agent()}")
webdriver_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=webdriver_service, options=chrome_options)

# Scrape car brands
url = "https://www.sgcarmart.com/new_cars/newcars_brand_landing.php"
driver.get(url)
time.sleep(1)

brands = []
brand_elements = driver.find_elements(By.XPATH, "//div[@id='rightside_content']//td/a")

for brand_element in brand_elements:
    brand_text = brand_element.text
    brand = brand_text.replace(" cars", "").strip()
    brands.append(brand)

# Define URLs
base_url = "https://www.sgcarmart.com/used_cars/listing.php"
params_petrol = "?ORD=MAK_ASC&ASL=1&RPG=10&DP2=&DP1=&AVL=2&OPC[]=0&FUE=Petrol&CTS[]=18&VTS[]=10&VTS[]=11&VTS[]=12&VTS[]=13&VTS[]=2&VTS[]=3&VTS[]=7&VTS[]=8&VTS[]=9&PR2=&PR1=&BRSR={}"
params_petrol_ev = "?ORD=MAK_ASC&ASL=1&RPG=10&DP2=&DP1=&AVL=2&OPC[]=0&FUE=Petrol-Electric&CTS[]=18&VTS[]=10&VTS[]=11&VTS[]=12&VTS[]=13&VTS[]=2&VTS[]=3&VTS[]=7&VTS[]=8&VTS[]=9&PR2=&PR1=&BRSR={}"
params_ev = "?ORD=MAK_ASC&ASL=1&RPG=10&DP2=&DP1=&AVL=2&OPC[]=0&FUE=Electric&CTS[]=18&CTS[]=25&VTS[]=10&VTS[]=11&VTS[]=12&VTS[]=13&VTS[]=2&VTS[]=3&VTS[]=7&VTS[]=8&VTS[]=9&PR2=&PR1=&BRSR={}"

params_list = [
    ('Petrol', params_petrol),
    ('Hybrid', params_petrol_ev),
    ('EV', params_ev)
]

# Scrape car detail page
def scrape_car_details(link, brands, max_retries=3):
    retries = 0
    while retries < max_retries:
        try:
            driver.get(link)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//td[contains(@class, 'label')]")))
            time.sleep(random.uniform(0.5, 0.8))  # Shorter random sleep to avoid being blocked
            
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')

            # Extract make and model
            make_model = soup.find('a', class_='nounderline globaltitle').text.strip()
            make = next((brand for brand in brands if brand in make_model), None)
            if make:
                model = make_model.replace(make, "").strip()
            else:
                make_model_split = make_model.split(" ", 1)
                make = make_model_split[0] if len(make_model_split) > 0 else "Unknown"
                model = make_model_split[1] if len(make_model_split) > 1 else ""
            
            # Initialize a dictionary to hold the car details
            data = {
                "Make": make,
                "Model": model,
                "Link": link
            }
            
            # Extract all relevant information dynamically
            info_pairs = soup.select('#carInfo .row_bg, #carInfo .even_row, #carInfo .row_bg1')
            for row in info_pairs:
                label_td = row.find('td', class_='label')
                if label_td:
                    label = label_td.get_text(strip=True)
                    value_td = row.find('td', class_='font_red') or row.find('td', valign='top', class_=None)
                    if value_td:
                        data[label] = value_td.get_text(strip=True)
            
            # Manually extracting remaining fields if not present in dynamic extraction as a failsafe
            specific_info = {
                "Price": soup.find(string="Price").find_next("strong").get_text(strip=True) if soup.find(string="Price") else "NIL",
                "Depreciation (SGD)": next((item.get_text(strip=True) for item in soup.find_all("td") if "/yr" in item.get_text()), "NIL"),
                "Registration Date": soup.find(string="Reg Date").find_next("td").get_text(strip=True) if soup.find(string="Reg Date") else "NIL",
                "Mileage (km)": soup.find(string="Mileage").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Mileage") else "NIL",
                "Road Tax": soup.find(string="Road Tax").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Road Tax") else "NIL",
                "Dereg Value": soup.find(string="Dereg Value").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Dereg Value") else "NIL",
                "OMV": soup.find(string="OMV").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="OMV") else "NIL",
                "COE": soup.find(string="COE").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="COE") else "NIL",
                "ARF": soup.find(string="ARF").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="ARF") else "NIL",
                "Power (bhp)": soup.find(string="Power").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Power") else "NIL",
                "Number of Owners": soup.find(string="No. of Owners").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="No. of Owners") else "NIL",
                "Engine Capacity": re.sub(r'[^\d]', '', soup.find('strong', string=re.compile('Engine Cap')).find_next("div", class_='row_info').get_text(strip=True)) if soup.find('strong', string=re.compile('Engine Cap')) else "NIL",
                "Vehicle Type": soup.find(string="Type of Vehicle").find_next("a").get_text(strip=True) if soup.find(string="Type of Vehicle") else "NIL"
            }

            # Extract "Duration of COE Left" from the "Registration Date" field if it exists
            reg_date_field = soup.find(string="Reg Date")
            if reg_date_field:
                reg_date_value = reg_date_field.find_next("td").get_text(strip=True)
                coe_left = reg_date_value.split('(')[-1].strip(')') if '(' in reg_date_value else "NIL"
                reg_date_value = reg_date_value.split('(')[0].strip() if '(' in reg_date_value else reg_date_value
            else:
                reg_date_value = "NIL"
                coe_left = "NIL"
            
            specific_info["Registration Date"] = reg_date_value
            specific_info["Duration of COE Left"] = coe_left
            
            data.update(specific_info)
            
            # Clean up extracted data
            data["Depreciation (SGD)"] = data["Depreciation (SGD)"].split('/yr')[0].strip()
            data["Road Tax"] = data["Road Tax"].split('/yr')[0].strip()
            data["Mileage (km)"] = data["Mileage (km)"].split(' ')[0].replace(',', '')
            
            # Convert monetary values to floats if possible, otherwise keep as string
            for key in ["Price", "Dereg Value", "OMV", "COE", "ARF"]:
                if data[key] != "NIL":
                    cleaned_value = re.sub(r'[^\d.]', '', data[key])
                    try:
                        data[key] = float(cleaned_value) if cleaned_value else "NIL"
                    except ValueError:
                        data[key] = cleaned_value

            # Convert number of owners to integer if possible
            if data["Number of Owners"] != "NIL":
                try:
                    data["Number of Owners"] = int(re.sub(r'[^\d]', '', data["Number of Owners"]))
                except ValueError:
                    data["Number of Owners"] = data["Number of Owners"]

            # Split Power into kW and bhp and store them as numbers
            power_kW, power_bhp = "NIL", "NIL"
            if data["Power (bhp)"] != "NIL":
                power_text = data["Power (bhp)"]
                power_match = re.search(r'(\d+(?:\.\d+)?)\s*kW\s*\((\d+)\s*bhp\)', power_text)
                if power_match:
                    power_kW, power_bhp = float(power_match.group(1)), int(power_match.group(2))
            data["Power (kW)"] = power_kW
            data["Power (bhp)"] = power_bhp

            # Convert Engine Capacity to an integer
            if data["Engine Capacity"] != "NIL":
                try:
                    data["Engine Capacity"] = int(data["Engine Capacity"])
                except ValueError:
                    data["Engine Capacity"] = "NIL"

            # Calculate COE category
            coe_category = "NIL"
            if category == "EV":
                # Only check power for EVs
                if data["Power (kW)"] != "NIL" and data["Power (bhp)"] != "NIL":
                    if data["Power (kW)"] < 110 and data["Power (bhp)"] < 147:
                        coe_category = "A"
                    else:
                        coe_category = "B"
            else:
                # Check both engine capacity and power for non-EVs
                if data["Engine Capacity"] != "NIL" and data["Power (kW)"] != "NIL" and data["Power (bhp)"] != "NIL":
                    if data["Engine Capacity"] < 1600 and data["Power (kW)"] < 97 and data["Power (bhp)"] < 130:
                        coe_category = "A"
                    else:
                        coe_category = "B"
            data["COE Category"] = coe_category
            
            return data

        except TimeoutException:
            retries += 1
            print(f"Retrying {retries}/{max_retries} for link: {link}")
            if retries == max_retries:
                print(f"Failed to load the page after {max_retries} attempts: {link}")
                return {
                    "Make": "NIL",
                    "Model": "NIL",
                    "Price": "NIL",
                    "Depreciation (SGD)": "NIL",
                    "Registration Date": "NIL",
                    "Duration of COE Left": "NIL",
                    "Mileage (km)": "NIL",
                    "Road Tax": "NIL",
                    "Dereg Value": "NIL",
                    "OMV": "NIL",
                    "COE": "NIL",
                    "ARF": "NIL",
                    "Power (bhp)": "NIL",
                    "Power (kW)": "NIL",
                    "Number of Owners": "NIL",
                    "Link": link,
                    "Engine Capacity": "NIL",
                    "Vehicle Type": "NIL",
                    "COE Category": "NIL"
                }

# Step 2: Scrape car links for the first page only
car_links_per_category = {category: [] for category, _ in params_list}

for category, params in params_list:
    url = base_url + params.format(0)
    
    # Load the page
    driver.get(url)
    
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@class, 'car-model-title')]")))
        time.sleep(random.uniform(1, 2))  # Short random sleep to avoid being blocked
        
        # Extract car links from the listings
        car_elements = driver.find_elements(By.XPATH, "//a[contains(@class, 'car-model-title')]")
        car_links = [element.get_attribute("href") for element in car_elements[:10]]  # Limit to 10 cars
        
        print(f"Found {len(car_links)} car listings on the page for category {category}.")
        
        car_links_per_category[category].extend(car_links)
    except TimeoutException:
        print(f"TimeoutException: No car listings found for category {category}.")
        continue

# Step 3: Scrape each car link for details
data_list_per_category = {category: [] for category, _ in params_list}

for category, car_links in car_links_per_category.items():
    for i, link in enumerate(car_links):  # Scrape all links
        car_data = scrape_car_details(link, brands)
        data_list_per_category[category].append(car_data)
        print(f"Scraped car details for link: {link}")

# Create a new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet created with the workbook
for category, data_list in data_list_per_category.items():
    df = pd.DataFrame(data_list)
    # Ensure that all expected columns are present in the DataFrame
    for col in ["Make", "Model", "Price", "Depreciation (SGD)", "Registration Date", "Duration of COE Left", "Mileage (km)", "Road Tax", "Dereg Value", "OMV", "COE", "ARF", "Power (bhp)", "Power (kW)", "Number of Owners", "Link", "Engine Capacity", "Vehicle Type", "COE Category"]:
        if col not in df.columns:
            df[col] = "NIL"
    df = df[["Make", "Model", "Price", "Depreciation (SGD)", "Registration Date", "Duration of COE Left", "Mileage (km)", "Road Tax", "Dereg Value", "OMV", "COE", "ARF", "Power (bhp)", "Power (kW)", "Number of Owners", "Link", "Engine Capacity", "Vehicle Type", "COE Category"]]
    
    # Add the data to a new sheet
    sheet_name = f"{category} Used Cars"
    ws = wb.create_sheet(title=sheet_name)
    
    # Write the DataFrame to the sheet
    for r_idx, row in enumerate(df.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col)

    # Apply currency formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=12):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=15, max_col=15):
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER

    # Auto-adjust column widths based on the longest cell content
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)  # Get the column letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook

timestamp = datetime.now().strftime('%d%m%y_%H%M')
file_name = f'temp_NEVC_Prices_Used_{timestamp}.xlsx'

wb.save(file_name)

print(f"Data has been saved to {file_name}")

# Close the driver
driver.quit()

            # Initialize a dictionary to hold the car details
            data = {
                "Make": make,
                "Model": model,
                "Link": link
            }
            
            # Extract all relevant information dynamically
            info_pairs = soup.select('#carInfo .row_bg, #carInfo .even_row, #carInfo .row_bg1')
            for row in info_pairs:
                label_td = row.find('td', class_='label')
                if label_td:
                    label = label_td.get_text(strip=True)
                    value_td = row.find('td', class_='font_red') or row.find('td', valign='top', class_=None)
                    if value_td:
                        data[label] = value_td.get_text(strip=True)
            
            # Manually extracting remaining fields if not present in dynamic extraction as a failsafe
            specific_info = {
                "Price": soup.find(string="Price").find_next("strong").get_text(strip=True) if soup.find(string="Price") else "NIL",
                "Depreciation (SGD)": next((item.get_text(strip=True) for item in soup.find_all("td") if "/yr" in item.get_text()), "NIL"),
                "Registration Date": soup.find(string="Reg Date").find_next("td").get_text(strip=True) if soup.find(string="Reg Date") else "NIL",
                "Mileage (km)": soup.find(string="Mileage").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Mileage") else "NIL",
                "Road Tax": soup.find(string="Road Tax").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Road Tax") else "NIL",
                "Dereg Value": soup.find(string="Dereg Value").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Dereg Value") else "NIL",
                "OMV": soup.find(string="OMV").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="OMV") else "NIL",
                "COE": soup.find(string="COE").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="COE") else "NIL",
                "ARF": soup.find(string="ARF").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="ARF") else "NIL",
                "Power (bhp)": soup.find(string="Power").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="Power") else "NIL",
                "Number of Owners": soup.find(string="No. of Owners").find_next("div", class_='row_info').get_text(strip=True) if soup.find(string="No. of Owners") else "NIL",
                "Engine Capacity": re.sub(r'[^\d]', '', soup.find('strong', string=re.compile('Engine Cap')).find_next("div", class_='row_info').get_text(strip=True)) if soup.find('strong', string=re.compile('Engine Cap')) else "NIL",
                "Vehicle Type": soup.find(string="Type of Vehicle").find_next("a").get_text(strip=True) if soup.find(string="Type of Vehicle") else "NIL"
            }

            # Extract "Duration of COE Left" from the "Registration Date" field if it exists
            reg_date_field = soup.find(string="Reg Date")
            if reg_date_field:
                reg_date_value = reg_date_field.find_next("td").get_text(strip=True)
                coe_left = reg_date_value.split('(')[-1].strip(')') if '(' in reg_date_value else "NIL"
                reg_date_value = reg_date_value.split('(')[0].strip() if '(' in reg_date_value else reg_date_value
            else:
                reg_date_value = "NIL"
                coe_left = "NIL"
            
            specific_info["Registration Date"] = reg_date_value
            specific_info["Duration of COE Left"] = coe_left
            
            data.update(specific_info)
            
            # Clean up extracted data
            data["Depreciation (SGD)"] = data["Depreciation (SGD)"].split('/yr')[0].strip()
            data["Road Tax"] = data["Road Tax"].split('/yr')[0].strip()
            data["Mileage (km)"] = data["Mileage (km)"].split(' ')[0].replace(',', '')
            
            # Convert monetary values to floats if possible, otherwise keep as string
            for key in ["Price", "Dereg Value", "OMV", "COE", "ARF"]:
                if data[key] != "NIL":
                    cleaned_value = re.sub(r'[^\d.]', '', data[key])
                    try:
                        data[key] = float(cleaned_value) if cleaned_value else "NIL"
                    except ValueError:
                        data[key] = cleaned_value

            # Convert number of owners to integer if possible
            if data["Number of Owners"] != "NIL":
                try:
                    data["Number of Owners"] = int(re.sub(r'[^\d]', '', data["Number of Owners"]))
                except ValueError:
                    data["Number of Owners"] = data["Number of Owners"]

            # Split Power into kW and bhp and store them as numbers
            power_kW, power_bhp = "NIL", "NIL"
            if data["Power (bhp)"] != "NIL":
                power_text = data["Power (bhp)"]
                power_match = re.search(r'(\d+(?:\.\d+)?)\s*kW\s*\((\d+)\s*bhp\)', power_text)
                if power_match:
                    power_kW, power_bhp = float(power_match.group(1)), int(power_match.group(2))
            data["Power (kW)"] = power_kW
            data["Power (bhp)"] = power_bhp

            # Convert Engine Capacity to an integer
            if data["Engine Capacity"] != "NIL":
                try:
                    data["Engine Capacity"] = int(data["Engine Capacity"])
                except ValueError:
                    data["Engine Capacity"] = "NIL"

            # Calculate COE category
            coe_category = "NIL"
            if category == "EV":
                # Only check power for EVs
                if data["Power (kW)"] != "NIL" and data["Power (bhp)"] != "NIL":
                    if data["Power (kW)"] < 110 and data["Power (bhp)"] < 147:
                        coe_category = "A"
                    else:
                        coe_category = "B"
            else:
                # Check both engine capacity and power for non-EVs
                if data["Engine Capacity"] != "NIL" and data["Power (kW)"] != "NIL" and data["Power (bhp)"] != "NIL":
                    if data["Engine Capacity"] < 1600 and data["Power (kW)"] < 97 and data["Power (bhp)"] < 130:
                        coe_category = "A"
                    else:
                        coe_category = "B"
            data["COE Category"] = coe_category
            
            return data

        except TimeoutException:
            retries += 1
            print(f"Retrying {retries}/{max_retries} for link: {link}")
            if retries == max_retries:
                print(f"Failed to load the page after {max_retries} attempts: {link}")
                return {
                    "Make": "NIL",
                    "Model": "NIL",
                    "Price": "NIL",
                    "Depreciation (SGD)": "NIL",
                    "Registration Date": "NIL",
                    "Duration of COE Left": "NIL",
                    "Mileage (km)": "NIL",
                    "Road Tax": "NIL",
                    "Dereg Value": "NIL",
                    "OMV": "NIL",
                    "COE": "NIL",
                    "ARF": "NIL",
                    "Power (bhp)": "NIL",
                    "Power (kW)": "NIL",
                    "Number of Owners": "NIL",
                    "Link": link,
                    "Engine Capacity": "NIL",
                    "Vehicle Type": "NIL",
                    "COE Category": "NIL"
                }

# Step 2: Scrape all car links first
car_links_per_category = {category: [] for category, _ in params_list}

for category, params in params_list:
    page = 0
    while True:
        start = page * 100
        url = base_url + params.format(start)
        
        # Load the page
        driver.get(url)
        
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@class, 'car-model-title')]")))
            time.sleep(random.uniform(1, 2))  # Short random sleep to avoid being blocked
            
            # Extract car links from the listings
            car_elements = driver.find_elements(By.XPATH, "//a[contains(@class, 'car-model-title')]")
            car_links = [element.get_attribute("href") for element in car_elements]
            
            print(f"Found {len(car_links)} car listings on the page for category {category}.")
            
            if not car_links:
                print(f"No more car listings found for category {category}. Stopping.")
                break
            
            car_links_per_category[category].extend(car_links)
            page += 1
        except TimeoutException:
            print(f"TimeoutException: No car listings found for category {category} on page {page}. Stopping.")
            break

    print(f"Total car links scraped for category {category}: {len(car_links_per_category[category])}")

# Step 3: Scrape each car link for details
data_list_per_category = {category: [] for category, _ in params_list}

for category, car_links in car_links_per_category.items():
    for i, link in enumerate(car_links):  # Scrape all links
        car_data = scrape_car_details(link, brands)
        data_list_per_category[category].append(car_data)
        print(f"Scraped car details for link: {link}")

# Create a new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet created with the workbook
for category, data_list in data_list_per_category.items():
    df = pd.DataFrame(data_list)
    # Ensure that all expected columns are present in the DataFrame
    for col in ["Make", "Model", "Price", "Depreciation (SGD)", "Registration Date", "Duration of COE Left", "Mileage (km)", "Road Tax", "Dereg Value", "OMV", "COE", "ARF", "Power (bhp)", "Power (kW)", "Number of Owners", "Link", "Engine Capacity", "Vehicle Type", "COE Category"]:
        if col not in df.columns:
            df[col] = "NIL"
    df = df[["Make", "Model", "Price", "Depreciation (SGD)", "Registration Date", "Duration of COE Left", "Mileage (km)", "Road Tax", "Dereg Value", "OMV", "COE", "ARF", "Power (bhp)", "Power (kW)", "Number of Owners", "Link", "Engine Capacity", "Vehicle Type", "COE Category"]]
    
    # Add the data to a new sheet
    sheet_name = f"{category} Used Cars"
    ws = wb.create_sheet(title=sheet_name)
    
    # Write the DataFrame to the sheet
    for r_idx, row in enumerate(df.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col)

    # Apply currency formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=12):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=15, max_col=15):
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER

    # Auto-adjust column widths based on the longest cell content
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)  # Get the column letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook

timestamp = datetime.now().strftime('%d%m%y_%H%M')
file_name = f'NEVC_Prices_Used_{timestamp}.xlsx'

wb.save(file_name)

print(f"Data has been saved to {file_name}")

# Close the driver
driver.quit()
