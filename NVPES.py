# © [2024] National Electric Vehicle Centre, Land Transport Authority.
# All rights reserved. This code, in full or in part, is the property of the Land Transport Authority.
# No part of this code may be disclosed, reproduced, or distributed without prior written permission.

# Last updated: 19/07/2024
# Ver. 2.0

# New Vehicle Price Extraction System (NVPES)

import time
import random
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Generate random user agents
def generate_random_user_agent():
    browsers = ['Mozilla/5.0', 'AppleWebKit/537.36', 'Chrome/90.0.4430.93', 'Safari/537.36', 'Firefox/89.0']
    platforms = ['Windows NT 10.0; Win64; x64', 'Macintosh; Intel Mac OS X 10_15_7', 'X11; Ubuntu; Linux x86_64']
    return f"{random.choice(browsers)} ({random.choice(platforms)}) {random.choice(browsers)}"

# Setup
chrome_options = Options()
chrome_options.add_argument("--headless")  
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument(f"user-agent={generate_random_user_agent()}")
webdriver_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=webdriver_service, options=chrome_options)

# Initialize lists to store the extracted data
makes = []
models = []
specs = []
prices = []
withCOE = []
coe_cat_list = []
price_with_coe = []
vehicle_types = []

# List to store commercial vehicle models
commercial_models = []

# Step 1: Scrape car brands
url = "https://www.sgcarmart.com/new_cars/newcars_brand_landing.php"
driver.get(url)
time.sleep(1)

brands = []
brand_elements = driver.find_elements(By.XPATH, "//div[@id='rightside_content']//td/a")

for brand_element in brand_elements:
    brand_text = brand_element.text
    brand = brand_text.replace(" cars", "").strip()
    brands.append(brand)

# Main extraction function
def extract_data(driver, vehicle_type, coe_category=None):
    # Capture all relevant tables including those with different background colors
    car_tables = driver.find_elements(By.XPATH, "//table[@width='100%' and (@bgcolor='#FFFFFF' or @bgcolor='#F6FDFF')]")
    print(f"Found {len(car_tables)} car listings on the page.")
    
    for table in car_tables:
        try:
            # Extract the model name
            model_elements = table.find_elements(By.XPATH, ".//a[contains(@href, 'newcars_overview.php?CarCode=')]/strong")
            for model_element in model_elements:
                model_name = model_element.text.strip()
                make = next((brand for brand in brands if brand in model_name), "NIL")
                model = model_name.replace(make, "").strip() if make != "NIL" else model_name
                
                # Extract the specifications and prices
                spec_elements = table.find_elements(By.XPATH, ".//label")
                price_elements = table.find_elements(By.XPATH, ".//td[contains(text(), '$')]")
                bhp_elements = table.find_elements(By.XPATH, ".//td[contains(text(), 'bhp')]")
                
                if not spec_elements or not price_elements or not bhp_elements:
                    continue
                
                for spec_element, price_element, bhp_element in zip(spec_elements, price_elements, bhp_elements):
                    specification = spec_element.text.strip()
                    price_text = price_element.text.strip()
                    
                    # Determine if the price includes COE
                    coe_included = 'Y' if '(w/o COE)' not in price_text else 'N'
                    
                    # Extract the main price and convert to number
                    if '\n' in price_text:
                        price_lines = price_text.split('\n')
                        main_price = price_lines[0].strip().replace('$', '').replace(',', '')
                    else:
                        main_price = price_text.split(' $')[0].replace('$', '').replace(',', '') if ' $' in price_text else price_text.replace('$', '').replace(',', '')
                    
                    main_price = float(main_price)  # Convert to float

                    # Extract bhp
                    bhp_text = bhp_element.text.strip().replace('bhp', '').strip()
                    bhp_value = int(bhp_text)
                    
                    # Determine COE category based on vehicle type
                    if vehicle_type in ['Petrol', 'Diesel', 'Petrol-Electric', 'Diesel-Electric']:
                        coe_cat = coe_category
                    elif vehicle_type == 'Electric':
                        coe_cat = 'A' if bhp_value <= 147 else 'B'
                    
                    print(f"Extracted make: {make}, model: {model}, specification: {specification}, price: {main_price}, COE: {coe_included}, bhp: {bhp_value}, COE Category: {coe_cat}")
                    
                    # Append the data to the lists
                    makes.append(make)
                    models.append(model)
                    specs.append(specification)
                    prices.append(main_price)
                    withCOE.append(coe_included)
                    coe_cat_list.append(coe_cat)
                    vehicle_types.append(vehicle_type)
        except Exception as e:
            print(f"Error extracting data: {e}")

# Define URL patterns for different vehicle types
url_patterns = {
    'Electric': "?VT=Electric&RPG=60",
    'Petrol': ["?FUE=p&DT=CoeA&ASL=1&RPG=60", "?FUE=p&DT=CoeB&ASL=1&RPG=60"],
    'Diesel': ["?FUE=d&DT=CoeA&ASL=1&RPG=60", "?FUE=d&DT=CoeB&ASL=1&RPG=60"],
    'Petrol-Electric': ["?FUE=r&DT=CoeA&ASL=1&RPG=60", "?FUE=r&DT=CoeB&ASL=1&RPG=60"],
    'Diesel-Electric': ["?FUE=i&DT=CoeA&ASL=1&RPG=60", "?FUE=i&DT=CoeB&ASL=1&RPG=60"]
}

base_url = "https://www.sgcarmart.com/new_cars/newcars_listing.php"

# Scrape all vehicle data with dynamic pagination
def scrape_vehicle_data(base_url, url_patterns):
    for vehicle_type, params_list in url_patterns.items():
        if isinstance(params_list, list):
            for params in params_list:
                page = 0
                coe_category = params.split('DT=Coe')[-1][0]  # Extract 'A' or 'B'
                while True:
                    start = page * 60
                    url = f"{base_url}{params}&BRSR={start}"
                    try:
                        driver.get(url)
                        time.sleep(5)
                        car_elements = driver.find_elements(By.XPATH, "//table[@width='100%' and (@bgcolor='#FFFFFF' or @bgcolor='#F6FDFF')]")
                        if not car_elements:
                            print(f"No data found for {vehicle_type} on page {page}. Stopping.")
                            break
                        extract_data(driver, vehicle_type, coe_category)
                        page += 1
                    except Exception as e:
                        print(f"Error scraping data for {vehicle_type} on page {page}: {e}")
                        break
        else:
            page = 0
            while True:
                start = page * 60
                url = f"{base_url}{params_list}&BRSR={start}"
                try:
                    driver.get(url)
                    time.sleep(5)
                    car_elements = driver.find_elements(By.XPATH, "//table[@width='100%' and (@bgcolor='#FFFFFF' or @bgcolor='#F6FDFF')]")
                    if not car_elements:
                        print(f"No data found for {vehicle_type} on page {page}. Stopping.")
                        break
                    extract_data(driver, vehicle_type)
                    page += 1
                except Exception as e:
                    print(f"Error scraping data for {vehicle_type} on page {page}: {e}")
                    break

# Scrape commercial cars
def extract_commercial_data(driver):
    page = 0
    while True:
        start = page * 60
        url = f"https://www.sgcarmart.com/new_cars/newcars_listing.php?BRSR={start}&FUE=&VTS%5B%5D=1&RPG=60"
        driver.get(url)
        time.sleep(5)
        car_tables = driver.find_elements(By.XPATH, "//table[@width='100%' and (@bgcolor='#FFFFFF' or @bgcolor='#F6FDFF')]")
        if not car_tables:
            print(f"No more commercial vehicle data found on page {page}. Stopping.")
            break
        for table in car_tables:
            try:
                model_elements = table.find_elements(By.XPATH, ".//a[contains(@href, 'newcars_overview.php?CarCode=')]/strong")
                for model_element in model_elements:
                    model_name = model_element.text.strip()
                    make = next((brand for brand in brands if brand in model_name), None)
                    if make:
                        model = model_name.replace(make, "").strip()
                    else:
                        model_elements_split = model_name.split(" ", 1)
                        make = model_elements_split[0] if len(model_elements_split) > 0 else "Unknown"
                        model = model_elements_split[1] if len(model_elements_split) > 1 else ""
                        
                    commercial_models.append(model)
                    
            except Exception as e:
                print(f"Error extracting commercial vehicle data: {e}")
        page += 1

# Scrape COE prices
def extract_coe_prices(driver):
    url = "https://www.motorist.sg/coe-results"
    driver.get(url)
    time.sleep(5)

    try:
        coe_month_year = driver.find_element(By.XPATH, "/html/body/main/div/div[1]/div/div[1]/div/div[1]/div[1]/div/h2/span[2]").text
        coe_bidding = driver.find_element(By.XPATH, "/html/body/main/div/div[1]/div/div[1]/div/div[1]/div[1]/div/p").text
        coe_label = f"{coe_month_year} {coe_bidding}"

        coe_price_a = driver.find_element(By.XPATH, "/html/body/main/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[2]/td[2]/p").text.replace('$', '').replace(',', '')
        coe_price_b = driver.find_element(By.XPATH, "/html/body/main/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[2]/td[3]/p").text.replace('$', '').replace(',', '')
        coe_price_c = driver.find_element(By.XPATH, "/html/body/main/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[2]/td[4]/p").text.replace('$', '').replace(',', '')
        
        return coe_label, float(coe_price_a), float(coe_price_b), float(coe_price_c)
    except Exception as e:
        print(f"Error extracting COE prices: {e}")
        return None, None, None, None

# Perform scraping
scrape_vehicle_data(base_url, url_patterns)

# Scrape commercial vehicles
extract_commercial_data(driver)

# Get COE prices
coe_label, coe_price_a, coe_price_b, coe_price_c = extract_coe_prices(driver)

# Update COE category to 'C' if model appears in both lists
for i, model in enumerate(models):
    if model in commercial_models:
        coe_cat_list[i] = 'C'

# Calculate 'Price with COE'
for i in range(len(prices)):
    if withCOE[i] == 'Y':
        price_with_coe.append(prices[i])
    else:
        if coe_cat_list[i] == 'A':
            price_with_coe.append(prices[i] + coe_price_a)
        elif coe_cat_list[i] == 'B':
            price_with_coe.append(prices[i] + coe_price_b)
        elif coe_cat_list[i] == 'C':
            price_with_coe.append(prices[i] + coe_price_c)

# Create DataFrame
data = {
    'Make': makes,
    'Model': models,
    'Specification': specs,
    'Price (From SGCarMart)': prices,
    'With COE': withCOE,
    'COE Category': coe_cat_list,
    'Price with COE (SGD)': price_with_coe,
    'Vehicle Type': vehicle_types
}
df = pd.DataFrame(data)

# File will be saved as 'NEVC_Prices_New.xlsx' with a new sheet named by the current date and time
timestamp = datetime.now().strftime('%d%m%y_%H%M')
file_name = 'NEVC_Prices_New.xlsx'
sheet_name = f"EV Prices {timestamp}"

# Check if the file exists
if os.path.exists(file_name):
    # Load the workbook and get the existing sheets
    book = load_workbook(file_name)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
else:
    # Create a new file
    df.to_excel(file_name, index=False, sheet_name=sheet_name)

# Load the workbook to apply formatting
book = load_workbook(file_name)
sheet = book[sheet_name]

# Apply currency formatting
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = '"$"#,##0.00'
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=7):
    for cell in row:
        cell.number_format = '"$"#,##0.00'

# Apply conditional formatting to 'With COE' column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
    for cell in row:
        if cell.value == 'Y':
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            cell.font = Font(color='006100')
        elif cell.value == 'N':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.font = Font(color='9C0006')

# Apply conditional formatting to 'COE Category' column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
    for cell in row:
        if cell.value == 'A':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            cell.font = Font(color='9C5700')
        elif cell.value == 'B':
            cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
            cell.font = Font(color='9C0006')
        elif cell.value == 'C':
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            cell.font = Font(color='006100')

# Add COE prices information
sheet['I1'] = f"COE Car Prices as of {coe_label}" 
sheet['I1'].alignment = Alignment(horizontal='center')
sheet['I1'].font = Font(bold=True)
sheet['I2'] = 'Cat A (SGD)'
sheet['J2'] = coe_price_a
sheet['J2'].number_format = '"$"#,##0.00'
sheet['I3'] = 'Cat B (SGD)'
sheet['J3'] = coe_price_b
sheet['J3'].number_format = '"$"#,##0.00'
sheet['I4'] = 'Cat C (SGD)'
sheet['J4'] = coe_price_c
sheet['J4'].number_format = '"$"#,##0.00'

# Auto-adjust column widths based on the longest cell content
for column_cells in sheet.columns:
    max_length = 0
    column_letter = column_cells[0].column_letter  # Get the column letter
    for cell in column_cells:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    sheet.column_dimensions[column_letter].width = adjusted_width

book.save(file_name)

print(f"Data has been saved to {file_name}")
